from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Form
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional, Dict
from pydantic import BaseModel
import os
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
from pathlib import Path
import uuid
from datetime import datetime

app = FastAPI(title="Consolidador Excel API")

# Configuración CORS para React
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173", "http://localhost:5174"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuración
UPLOAD_FOLDER = "uploads"
TEMPLATE_FOLDER = "templates"
RESULTS_FOLDER = "results"
ALLOWED_EXTENSIONS = {'xlsm', 'xlsx'}

# Crear directorios
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TEMPLATE_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)

# Estado global de la aplicación
app_state = {
    "template_path": None,
    "template_name": None,
    "template_id": None,
    "sheet_names": [],
    "uploaded_files": {},  # {sessionId: [file_paths]}
    "tasks": {}  # {taskId: task_info}
}

# Modelos Pydantic
class TemplateUploadResponse(BaseModel):
    template_id: str
    template_name: str
    sheet_names: List[str]

class ConsolidateProcessRequest(BaseModel):
    session_id: str
    excluded_sheets: Optional[List[str]] = []

class TaskStatusResponse(BaseModel):
    task_id: str
    status: str  # "processing", "completed", "error"
    progress: int
    current_file: str
    status_message: str
    result_id: Optional[str] = None
    error: Optional[str] = None

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def update_task_progress(task_id: str, progress: int, current_file: str = "", message: str = ""):
    if task_id in app_state["tasks"]:
        app_state["tasks"][task_id].update({
            "progress": progress,
            "current_file": current_file,
            "status_message": message
        })

def mark_task_complete(task_id: str, result_id: str):
    if task_id in app_state["tasks"]:
        app_state["tasks"][task_id].update({
            "status": "completed",
            "progress": 100,
            "result_id": result_id
        })

def mark_task_error(task_id: str, error_msg: str):
    if task_id in app_state["tasks"]:
        app_state["tasks"][task_id].update({
            "status": "error",
            "error": error_msg
        })

def consolidate_xlsm_files(task_id: str, template_path: str, file_paths: List[str], 
                          output_path: str, included_sheets: List[str], result_id: str):
    """Función que realiza la consolidación de archivos Excel"""
    try:
        update_task_progress(task_id, 5, "Cargando plantilla", "Iniciando proceso...")
        wb_plantilla = openpyxl.load_workbook(template_path, keep_vba=True)
        
        hojas = [hoja for hoja in included_sheets if hoja in wb_plantilla.sheetnames]
        
        # Identificar celdas con fórmulas
        formula_cells = set()
        for hoja in hojas:
            ws = wb_plantilla[hoja]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        coord = f"{get_column_letter(cell.column)}{cell.row}"
                        formula_cells.add(f"{hoja}||{coord}")
        
        # Diccionario para acumular sumas
        sumas = {hoja: {} for hoja in hojas}
        total_files = len(file_paths)
        
        # Procesar cada archivo
        for i, file_path in enumerate(file_paths):
            try:
                filename = os.path.basename(file_path)
                update_task_progress(
                    task_id,
                    int(5 + (i + 1) / total_files * 85), 
                    filename, 
                    f"Procesando archivo {i + 1} de {total_files}"
                )
                
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                for hoja in hojas:
                    if hoja not in wb.sheetnames:
                        continue
                    
                    ws = wb[hoja]
                    for row in ws.iter_rows():
                        for cell in row:
                            coord = f"{get_column_letter(cell.column)}{cell.row}"
                            cell_key = f"{hoja}||{coord}"
                            
                            if isinstance(cell.value, (int, float)) and cell_key not in formula_cells:
                                if coord in sumas[hoja]:
                                    sumas[hoja][coord] += cell.value
                                else:
                                    sumas[hoja][coord] = cell.value
                
                wb.close()
                
            except Exception as e:
                print(f"Error procesando {file_path}: {e}")
                continue
        
        # Aplicar sumas a la plantilla
        update_task_progress(task_id, 95, "Generando resultado", "Aplicando sumas a la plantilla...")
        for hoja in hojas:
            if hoja not in wb_plantilla.sheetnames:
                continue
            
            ws = wb_plantilla[hoja]
            for coord, valor in sumas[hoja].items():
                cell_key = f"{hoja}||{coord}"
                if cell_key not in formula_cells:
                    try:
                        ws[coord].value = valor
                    except:
                        continue
        
        # Guardar con la extensión correcta
        if template_path.endswith('.xlsm'):
            output_path = output_path.replace('.xlsx', '.xlsm')
        
        wb_plantilla.save(output_path)
        wb_plantilla.close()
        
        update_task_progress(task_id, 100, "Completado", "Archivo consolidado listo para descargar")
        mark_task_complete(task_id, result_id)
        
        # Limpiar archivos temporales subidos
        for file_path in file_paths:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except:
                pass
        
        return True
    
    except Exception as e:
        print(f"Error en consolidación: {e}")
        mark_task_error(task_id, f"Error durante la consolidación: {str(e)}")
        return False


# ==================== ENDPOINTS ====================

@app.get("/")
async def root():
    return {
        "message": "API Consolidador Excel",
        "version": "2.0",
        "endpoints": {
            "template_download": "GET /api/template/download",
            "template_upload": "POST /api/template/upload",
            "consolidate_upload": "POST /api/consolidate/upload",
            "consolidate_process": "POST /api/consolidate/process",
            "consolidate_status": "GET /api/consolidate/status/{taskId}",
            "consolidate_download": "GET /api/consolidate/download/{resultId}"
        }
    }

# ==================== TEMPLATE ENDPOINTS ====================

@app.get("/api/template/download")
async def download_template():
    """
    Descarga el archivo SA_26_V1.1.xlsm ubicado en la raíz del proyecto.
    """
    # Nombre del archivo fijo en la raíz
    root_file_name = "SA_26_V1.1.xlsm"
    root_file_path = os.path.join(os.getcwd(), root_file_name)

    # 1. Intentar descargar el archivo de la raíz primero
    if os.path.exists(root_file_path):
        return FileResponse(
            path=root_file_path,
            filename=root_file_name,
            # Media type específico para archivos Excel con Macros (.xlsm)
            media_type='application/vnd.ms-excel.sheet.macroEnabled.12'
        )
    
    # 2. Si no existe en la raíz, intentar con lo que esté en app_state (tu lógica original)
    if app_state["template_path"] and os.path.exists(app_state["template_path"]):
        return FileResponse(
            path=app_state["template_path"],
            filename=app_state["template_name"],
            media_type='application/vnd.ms-excel.sheet.macroEnabled.12'
        )
    
    # 3. Si nada funciona, error
    raise HTTPException(
        status_code=404, 
        detail=f"Archivo {root_file_name} no encontrado en la raíz del proyecto."
    )


@app.post("/api/template/upload", response_model=TemplateUploadResponse)
async def upload_template(template: UploadFile = File(...)):
    """
    POST /api/template/upload
    Sube la plantilla maestra y retorna las hojas disponibles
    
    Returns:
    - template_id: ID único de la plantilla
    - template_name: Nombre del archivo
    - sheet_names: Lista de todas las hojas disponibles
    """
    
    # Validar tipo de archivo
    if not allowed_file(template.filename):
        raise HTTPException(
            status_code=400, 
            detail="Tipo de archivo no permitido. Use archivos .xlsx o .xlsm"
        )
    
    try:
        # Limpiar plantilla anterior si existe
        if app_state["template_path"] and os.path.exists(app_state["template_path"]):
            try:
                os.remove(app_state["template_path"])
            except:
                pass
        
        # Generar ID único para la plantilla
        template_id = uuid.uuid4().hex
        
        # Guardar nueva plantilla
        template_filename = f"template_{template_id}_{template.filename}"
        template_path = os.path.join(TEMPLATE_FOLDER, template_filename)
        
        with open(template_path, "wb") as buffer:
            shutil.copyfileobj(template.file, buffer)
        
        # Leer nombres de hojas
        wb = openpyxl.load_workbook(template_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # Actualizar estado
        app_state["template_path"] = template_path
        app_state["template_name"] = template.filename
        app_state["template_id"] = template_id
        app_state["sheet_names"] = sheet_names
        
        return TemplateUploadResponse(
            template_id=template_id,
            template_name=template.filename,
            sheet_names=sheet_names
        )
    
    except Exception as e:
        if 'template_path' in locals() and os.path.exists(template_path):
            os.remove(template_path)
        raise HTTPException(
            status_code=500, 
            detail=f"Error al procesar la plantilla: {str(e)}"
        )


# ==================== CONSOLIDATE ENDPOINTS ====================

@app.post("/api/consolidate/upload")
async def upload_files_to_consolidate(files: List[UploadFile] = File(...)):
    """
    POST /api/consolidate/upload
    Sube los archivos que serán consolidados
    
    Returns:
    - session_id: ID de sesión para referenciar estos archivos
    - files_count: Cantidad de archivos subidos
    - files: Lista de nombres de archivos
    """
    
    # Generar ID de sesión único
    session_id = uuid.uuid4().hex
    
    # Guardar archivos
    saved_files = []
    file_names = []
    
    for file in files:
        if file and allowed_file(file.filename):
            file_id = uuid.uuid4().hex[:8]
            file_path = os.path.join(UPLOAD_FOLDER, f"{session_id}_{file_id}_{file.filename}")
            
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            saved_files.append(file_path)
            file_names.append(file.filename)
    
    if not saved_files:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron archivos válidos para consolidar"
        )
    
    # Guardar en estado
    app_state["uploaded_files"][session_id] = saved_files
    
    return {
        "session_id": session_id,
        "files_count": len(saved_files),
        "files": file_names,
        "message": "Archivos subidos correctamente. Use POST /api/consolidate/process para iniciar la consolidación."
    }


@app.post("/api/consolidate/process")
async def process_consolidation(
    background_tasks: BackgroundTasks,
    session_id: str = Form(...),
    excluded_sheets: Optional[str] = Form(None)
):
    """
    POST /api/consolidate/process
    Inicia el proceso de consolidación
    
    Parameters:
    - session_id: ID de sesión de los archivos subidos
    - excluded_sheets: Hojas a excluir separadas por coma (opcional)
    
    Returns:
    - task_id: ID único de la tarea para consultar el estado
    - message: Mensaje de confirmación
    """
    
    # Validar que exista plantilla
    if not app_state["template_path"] or not os.path.exists(app_state["template_path"]):
        raise HTTPException(
            status_code=400,
            detail="No hay plantilla cargada. Primero sube una plantilla usando POST /api/template/upload"
        )
    
    # Validar que existan archivos para el session_id
    if session_id not in app_state["uploaded_files"]:
        raise HTTPException(
            status_code=404,
            detail=f"No se encontraron archivos para el session_id: {session_id}"
        )
    
    file_paths = app_state["uploaded_files"][session_id]
    
    # Procesar hojas excluidas
    excluded_list = []
    if excluded_sheets:
        excluded_list = [s.strip() for s in excluded_sheets.split(',') if s.strip()]
    
    included_sheets = [
        sheet for sheet in app_state["sheet_names"] 
        if sheet not in excluded_list
    ]
    
    # Generar IDs únicos
    task_id = uuid.uuid4().hex
    result_id = uuid.uuid4().hex
    
    # Preparar salida
    output_filename = f'REM_Consolidado_{result_id}_{app_state["template_name"]}'
    output_path = os.path.join(RESULTS_FOLDER, output_filename)
    
    # Crear tarea
    app_state["tasks"][task_id] = {
        "task_id": task_id,
        "status": "processing",
        "progress": 0,
        "current_file": "",
        "status_message": "Iniciando consolidación...",
        "result_id": None,
        "result_filename": output_filename,
        "error": None,
        "created_at": datetime.now().isoformat()
    }
    
    # Ejecutar consolidación en background
    background_tasks.add_task(
        consolidate_xlsm_files,
        task_id,
        app_state["template_path"],
        file_paths,
        output_path,
        included_sheets,
        result_id
    )
    
    return {
        "task_id": task_id,
        "message": "Consolidación iniciada",
        "status_url": f"/api/consolidate/status/{task_id}",
        "included_sheets": included_sheets,
        "excluded_sheets": excluded_list
    }


@app.get("/api/consolidate/status/{task_id}", response_model=TaskStatusResponse)
async def get_consolidation_status(task_id: str):
    """
    GET /api/consolidate/status/{taskId}
    Obtiene el estado actual de una tarea de consolidación
    
    Returns:
    - task_id: ID de la tarea
    - status: Estado actual ("processing", "completed", "error")
    - progress: Porcentaje de progreso (0-100)
    - current_file: Archivo que se está procesando
    - status_message: Mensaje descriptivo del estado
    - result_id: ID para descargar el resultado (solo cuando status="completed")
    - error: Mensaje de error (solo cuando status="error")
    """
    
    if task_id not in app_state["tasks"]:
        raise HTTPException(
            status_code=404,
            detail=f"No se encontró la tarea con ID: {task_id}"
        )
    
    task_info = app_state["tasks"][task_id]
    
    return TaskStatusResponse(
        task_id=task_info["task_id"],
        status=task_info["status"],
        progress=task_info["progress"],
        current_file=task_info["current_file"],
        status_message=task_info["status_message"],
        result_id=task_info["result_id"],
        error=task_info["error"]
    )


@app.get("/api/consolidate/download/{result_id}")
async def download_consolidated_file(result_id: str):
    """
    GET /api/consolidate/download/{resultId}
    Descarga el archivo consolidado
    
    Parameters:
    - result_id: ID del resultado generado
    """
    
    # Buscar el archivo en la carpeta de resultados
    result_files = [f for f in os.listdir(RESULTS_FOLDER) if result_id in f]
    
    if not result_files:
        raise HTTPException(
            status_code=404,
            detail=f"No se encontró el resultado con ID: {result_id}"
        )
    
    file_path = os.path.join(RESULTS_FOLDER, result_files[0])
    
    if not os.path.exists(file_path):
        raise HTTPException(
            status_code=404,
            detail="Archivo no encontrado o ya fue eliminado"
        )
    
    return FileResponse(
        path=file_path,
        filename=result_files[0],
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.post("/api/consolidate/validate")
async def validate_template_version(file: UploadFile = File(...)):
    try:
        # 1. Guardar temporalmente
        temp_path = os.path.join(UPLOAD_FOLDER, f"val_{uuid.uuid4().hex}_{file.filename}")
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # 2. Cargar con data_only=True para leer el texto final
        # keep_vba=True es opcional pero recomendado para archivos .xlsm
        wb = openpyxl.load_workbook(temp_path, data_only=True, keep_vba=True)
        
        # 3. Acceder específicamente a la primera pestaña por nombre o índice
        # Tu archivo tiene la pestaña "IDENTIFICACIÓN" como la primera
        sheet = wb.worksheets[0] 
        
        # 4. Leer A9 (Celda maestra de la combinación A9:B9)
        valor_celda = sheet["A9"].value
        version_encontrada = str(valor_celda).strip() if valor_celda else ""

        wb.close()
        os.remove(temp_path)

        # 5. Comparación exacta
        VERSION_ESPERADA = "Versión 1.1: Febrero 2026"
        
        if version_encontrada == VERSION_ESPERADA:
            return {
                "status": "success",
                "message": "Validación exitosa: La plantilla es la correcta.",
                "version": version_encontrada
            }
        else:
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "detail": "La versión del documento no coincide.",
                    "encontrado": version_encontrada,
                    "esperado": VERSION_ESPERADA
                }
            )

    except Exception as e:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
# ==================== UTILITY ENDPOINTS ====================

@app.get("/api/health")
async def health_check():
    """Verifica el estado de la API"""
    return {
        "status": "healthy",
        "template_loaded": app_state["template_path"] is not None,
        "template_name": app_state["template_name"],
        "active_sessions": len(app_state["uploaded_files"]),
        "active_tasks": len([t for t in app_state["tasks"].values() if t["status"] == "processing"])
    }


@app.delete("/api/cleanup")
async def cleanup_old_files():
    """Limpia archivos antiguos y sesiones completadas"""
    
    cleaned = {
        "templates": 0,
        "uploads": 0,
        "results": 0,
        "sessions": 0,
        "tasks": 0
    }
    
    # Limpiar tareas completadas/con error (más de 1 hora)
    from datetime import datetime, timedelta
    cutoff_time = datetime.now() - timedelta(hours=1)
    
    tasks_to_remove = []
    for task_id, task_info in app_state["tasks"].items():
        if task_info["status"] in ["completed", "error"]:
            task_time = datetime.fromisoformat(task_info["created_at"])
            if task_time < cutoff_time:
                tasks_to_remove.append(task_id)
    
    for task_id in tasks_to_remove:
        del app_state["tasks"][task_id]
        cleaned["tasks"] += 1
    
    # Limpiar sesiones sin tareas asociadas
    sessions_to_remove = []
    for session_id in app_state["uploaded_files"].keys():
        # Si no hay tareas activas para esta sesión, limpiar
        has_active_task = any(
            session_id in str(task_info) 
            for task_info in app_state["tasks"].values()
        )
        if not has_active_task:
            sessions_to_remove.append(session_id)
    
    for session_id in sessions_to_remove:
        del app_state["uploaded_files"][session_id]
        cleaned["sessions"] += 1
    
    return {
        "success": True,
        "message": "Limpieza completada",
        "cleaned": cleaned
    }


@app.delete("/api/reset")
async def reset_state():
    """Reinicia completamente el estado de la aplicación"""
    
    # Limpiar plantilla
    if app_state["template_path"] and os.path.exists(app_state["template_path"]):
        try:
            os.remove(app_state["template_path"])
        except:
            pass
    
    # Limpiar archivos subidos
    for session_files in app_state["uploaded_files"].values():
        for file_path in session_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except:
                pass
    
    # Resetear estado
    app_state["template_path"] = None
    app_state["template_name"] = None
    app_state["template_id"] = None
    app_state["sheet_names"] = []
    app_state["uploaded_files"] = {}
    app_state["tasks"] = {}
    
    return {
        "success": True,
        "message": "Estado reiniciado completamente"
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
